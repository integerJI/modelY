from django.shortcuts import render
import io
from django.http import HttpResponse
from xlsxwriter.workbook import Workbook
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def index(request):
    return render(request, 'index.html')

def download(request):
    sheetPath = request.GET['sheetPath']

    scope = [
        'https://spreadsheets.google.com/feeds'
        , 'https://www.googleapis.com/auth/drive'
    ]

    json_file_name = 'calm-magpie-382913-e094e6764e08.json'

    credentials = ServiceAccountCredentials.from_json_keyfile_name(json_file_name, scope)
    gc = gspread.authorize(credentials)

    spreadsheet_url = sheetPath

    # 스프레드시트 문서 가져오기 
    doc = gc.open_by_url(spreadsheet_url)

    # 시트 선택하기
    worksheet = doc.worksheet('판매현황')

    # 열 데이터 가져오기
    column_data = worksheet.col_values(3)
    order_date_data = worksheet.col_values(11)
    company_code = worksheet.col_values(5)
    order_number = worksheet.col_values(14)
    brand_name = worksheet.col_values(17)

    product_name1 = worksheet.col_values(18)
    product_name2 = worksheet.col_values(19)

    ea_name = worksheet.col_values(20)
    color_name = worksheet.col_values(21)
    size_name = worksheet.col_values(22)
    retail_price = worksheet.col_values(32)
    supply_price = worksheet.col_values(33)
    sale_rate = worksheet.col_values(40)

    num_cells = len(company_code)

    for i in range(num_cells):
        if i >= len(company_code):  # 마지막 셀까지 처리한 경우 break
            break
        if company_code[i] == '':  # 빈 문자열이면 새로운 값 추가
            company_code.append('')

    num_sale_cells = len(sale_rate)

    for i in range(num_sale_cells):
        if i >= len(sale_rate):  # 마지막 셀까지 처리한 경우 break
            break
        if sale_rate[i] == '':  # 빈 문자열이면 새로운 값 추가
            sale_rate.append('')

    # 2차원 리스트로 변환
    column_data_2d = [[col] for col in column_data]
    order_date_data_2d = [[col] for col in order_date_data]
    company_code_2d = [[col] for col in company_code]
    order_number_2d = [[col] for col in order_number]
    brand_name_2d = [[col] for col in brand_name]
    product_name1_2d = [[col] for col in product_name1]
    product_name2_2d = [[col] for col in product_name2]
    ea_name_2d = [[col] for col in ea_name]
    color_name_2d = [[col] for col in color_name]
    size_name_2d = [[col] for col in size_name]
    retail_price_2d = [[col] for col in retail_price]
    supply_price_2d = [[col] for col in supply_price]
    sale_rate_2d = [[col] for col in sale_rate]

    # 엑셀 파일 생성
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active

    # 첫 번째 행에 값 넣기
    worksheet.append(['주문일', '출고일', '코드번호', '주문번호', '브랜드', '상품명1', '상품명2', '수량', '컬러', '사이즈', '소매가', '공급률', '공급가', '주문액', '세일', '택배', '송장번호', '비고'])

    # 스타일 지정
    font = Font(color='FFFFFF')
    fill = PatternFill(patternType='solid', fgColor='000000')
    align = Alignment(horizontal='center', vertical='center', wrapText=True)

    # 스타일 적용
    for cell in worksheet[1]:
        cell.font = font
        cell.fill = fill
        cell.alignment = align

    # 첫 번째 열의 너비를 20으로 설정
    worksheet.column_dimensions[get_column_letter(1)].width = 15
    worksheet.column_dimensions[get_column_letter(2)].width = 15
    worksheet.column_dimensions[get_column_letter(3)].width = 20
    worksheet.column_dimensions[get_column_letter(4)].width = 20
    worksheet.column_dimensions[get_column_letter(5)].width = 15
    worksheet.column_dimensions[get_column_letter(6)].width = 20
    worksheet.column_dimensions[get_column_letter(7)].width = 30
    worksheet.column_dimensions[get_column_letter(8)].width = 10

    # 첫 번째 행(리스트) 삭제
    del column_data_2d[0]
    del order_date_data_2d[0]
    del company_code_2d[0]
    del order_number_2d[0]
    del brand_name_2d[0]
    del product_name1_2d[0]
    del product_name2_2d[0]
    del ea_name_2d[0]
    del color_name_2d[0]
    del size_name_2d[0]
    del retail_price_2d[0]
    del supply_price_2d[0]
    del sale_rate_2d[0]

    # 시트 데이터를 엑셀 파일에 저장
    for i, col in enumerate(column_data_2d):
        worksheet.cell(row=i+2, column=1, value=order_date_data_2d[i][0])
        worksheet.cell(row=i+2, column=3, value=company_code_2d[i][0])
        worksheet.cell(row=i+2, column=4, value=order_number_2d[i][0])
        worksheet.cell(row=i+2, column=5, value=brand_name_2d[i][0])
        worksheet.cell(row=i+2, column=6, value=product_name1_2d[i][0])
        worksheet.cell(row=i+2, column=7, value=product_name2_2d[i][0])
        worksheet.cell(row=i+2, column=8, value=ea_name_2d[i][0])
        worksheet.cell(row=i+2, column=9, value=color_name_2d[i][0])
        worksheet.cell(row=i+2, column=10, value=size_name_2d[i][0])
        worksheet.cell(row=i+2, column=11, value=retail_price_2d[i][0])
        worksheet.cell(row=i+2, column=13, value=supply_price_2d[i][0])
        worksheet.cell(row=i+2, column=15, value=sale_rate_2d[i][0])
    workbook.save(output)

    # response 생성
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=새로운 엑셀 파일명.xlsx'
    response.write(output.getvalue())

    return response
